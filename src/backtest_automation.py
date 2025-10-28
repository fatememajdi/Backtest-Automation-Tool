from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles import Alignment, Border, Side
import re
import numpy as np
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from machwise import Wise
from dotenv import load_dotenv
import os
from dotenv import load_dotenv
import os

env_file = os.path.join(os.getcwd(), ".env")
load_dotenv(env_file)

API_KEY = os.getenv("API_KEY")

def GetMachineInformation(lst_name, SelectedDate):
    """
    Fetch machine-specific information from the API and save it locally as JSON.

    Args:
        lst_name (list): List of machine tags to query.
        SelectedDate (str): Date string to query machine data for.

    Saves:
        JSON file at '/Users/fateme/Desktop/CeeNous/Backtest Automation/data/machin_info.json'
    """

    headers = {
        "api-key": API_KEY
    }
    input_data = {
        'lst_name': lst_name,
        'server_name': 'https://api.seenous.net',
        'rpm_variable': {},
        'SelectedDate': SelectedDate
    }

    res = requests.post(
        url=f"http://20.108.64.195:4000/pointinformations", json=input_data, headers=headers)

    print("Status code:", res.status_code)

    with open(f'/Users/fateme/Desktop/CeeNous/Backtest Automation/data/machin_info.json', "w", encoding="utf-8") as f:
        json.dump(res.json(), f, ensure_ascii=False, indent=2)


def GetCompanyData(company_id, domain, company_domain):
    """
    Fetch all diagnostic information for a given company from a remote API.

    Args:
        company_id (str): Company identifier.
        domain (str): Server domain to send the request to.
        company_domain (str): Specific company domain name used for saving results.

    Returns:
        dict: JSON response containing all diagnostic information, or None if failed.

    Side Effects:
        Saves response as JSON locally at:
        '/Users/fateme/Desktop/CeeNous/Backtest Automation/data/response_{company_domain}.json'
    """

    url = f"http://{domain}:4000/alldiagnosisinformations"
    headers = {
        "api-key": API_KEY
    }

    input_data = {
        'CompanyId': company_id,
        'SelectedDate': '',
        "domain": company_domain
    }

    try:
        res = requests.post(url=url, json=input_data,
                            headers=headers, timeout=30)
        print("Status code:", res.status_code)

        if res.status_code == 200:
            try:
                return res.json()
            except json.JSONDecodeError:
                print("Invalid JSON response")
                return None
        else:
            print(f"Request failed: {res.status_code} - {res.text}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error in GetCompanyData: {e}")
        return None


def _extract_chart_key(chartplot: str, tag: str):
    """
    Extract a standardized chart key from a chartplot filename.

    Args:
        chartplot (str): The chart plot filename string.
        tag (str): Machine tag to remove from the filename prefix.

    Returns:
        str: Extracted chart key (e.g., "M-1-H") or empty string if not found.
    """
    if not chartplot:
        return ""
    chartplot = chartplot.strip()

    if chartplot.startswith(tag):
        chartplot = chartplot[len(tag):].lstrip('-')

    match = re.search(r'([A-Za-z]+)-(\d+)-([A-Za-z])', chartplot)
    if match:
        return "-".join(match.groups()).upper()
    return ""


def get_spare_full_name(letter):
    """
    Convert a single-letter spare part code to its full name.

    Args:
        letter (str): Spare part code (M, G, P, F).

    Returns:
        str: Full name of the spare part.
    """
    mapping = {
        'M': 'Electromotor',
        'G': 'Gearbox',
        'P': 'Pinion',
        'F': 'Fan'
    }
    return mapping.get(letter.upper(), letter)


def RPM_range(rpm):
    """
    Categorize RPM value into a range label.

    Args:
        rpm (float): RPM value.

    Returns:
        str: One of 'low_rpm', 'mid_rpm', 'high_rpm'.
    """
    if rpm > 300:
        return 'high_rpm'
    elif rpm < 120:
        return 'low_rpm'
    else:
        return 'mid_rpm'


def _safe_float(x):
    """
    Safely convert a value to float, returning NaN for invalid values.

    Args:
        x: Input value to convert.

    Returns:
        float: Converted float value or np.nan if conversion fails.
    """
    try:
        if x in [None, "", "-", "nan", "NaN"]:
            return np.nan
        return float(x)
    except:
        return np.nan


def _norm(x):
    """
    Normalize a value to a stripped string, returning empty string if None.

    Args:
        x: Input value.

    Returns:
        str: Normalized string.
    """
    return "" if x is None else str(x).strip()


def find_post(post_json, tag, date, spare, point, failure_pre=None):
    """
    Find a post-test record corresponding to a specific pre-test record.

    Args:
        post_json (dict): Post-test data structured as JSON.
        tag (str): Machine tag.
        date (str): Date string.
        spare (str): Spare part name/code.
        point (str): Measurement point.
        failure_pre (str, optional): Pre-test failure description.

    Returns:
        dict or None: Matching post-test record or None if not found.
    """
    tag, date, spare, point = map(_norm, [tag, date, spare, point])
    if tag not in post_json or date not in post_json[tag]:
        return None

    candidates = post_json[tag][date]
    if not candidates:
        return None

    if failure_pre and failure_pre.strip().lower() == "no failure detected":
        for r in candidates:
            if spare.lower() in _norm(r.get("SparePart", "")).lower():
                return r
    else:
        for r in candidates:
            if (
                _norm(r.get("SparePart", "")).lower() == spare.lower()
                and _norm(r.get("Point", "")).lower() == point.lower()
            ):
                return r
    return None


def _generate_base_rows(tag, date, info):
    """
    Generate baseline rows for all machine parts with default 'Healthy' fault status.

    Args:
        tag (str): Machine tag.
        date (str): Selected date.
        info (Wise): Machine information object containing parts and RPM data.

    Returns:
        list: List of rows with default faults for each machine part.
    """
    rows = []

    for fault in ['Misalignment', 'Structural Looseness', 'Unbalance', 'Gear Failure']:
        rpm_val = RPM_range(info.rpm['1']['value'])
        parts = info.machine_part.keys()

        # Electromotor
        if 'M' in parts and fault in ['Misalignment', 'Structural looseness', 'Unbalance']:
            rows.append([tag, 'Electromotor', "-", rpm_val, date, fault,
                         'Healthy', '-', True, fault, '-', '', True, "", '0', True, "", ''])

        # Fan
        if 'F' in parts and fault in ['Structural looseness', 'Unbalance']:
            rows.append([tag, 'Fan', "-", rpm_val, date, fault,
                         'Healthy', '-', True, fault, '-', '', True, "", '0', True, "", ''])

        # Pump
        if 'Pump' in parts and fault == 'Unbalance':
            rows.append([tag, 'Pump', "-", rpm_val, date, fault,
                         'Healthy', '-', True, fault, '-', '', True, "", '0', True, "", ''])

        # Gear / Pinion
        if fault == 'Gear failure':
            if 'P' in parts:
                rows.append([tag, 'Pinion', "-", rpm_val, date, fault,
                             'Healthy', '-', True, fault, '-', '', True, "", '0', True, "", ''])
            if 'G' in parts:
                rows.append([tag, 'Gearbox', "-", rpm_val, date, fault,
                             'Healthy', '-', True, fault, '-', '', True, "", '0', True, "", ''])

    # Bearings & looseness
    for spare_part, points in info.machine_part.items():
        for point in points:
            rpm_val = RPM_range(info.rpm[point]['value'])
            for fault in ['Bearing Failure', 'Mechanical Looseness']:
                rows.append([
                    tag, get_spare_full_name(spare_part), point, rpm_val,
                    date, fault, 'Healthy', '-', True, fault, '-', '',
                    True, "", '0', True, "", ''
                ])

    return rows


def _update_with_records(new_rows, records, tag, date, post_json):
    """
    Update baseline rows with actual pre/post diagnostic records.

    Args:
        new_rows (list): Generated baseline rows.
        records (list): List of pre-test records.
        tag (str): Machine tag.
        date (str): Selected date.
        post_json (dict): Post-test JSON data.

    Returns:
        list: Updated rows including pre/post test data and validation.
    """
    for rec in records:
        machine = _norm(rec.get("Tag", tag))
        part = _norm(rec.get("SparePart"))
        point = _norm(rec.get("Point"))
        failure = _norm(rec.get("FailureDiagnosis"))
        date_ = _norm(rec.get("SelectedDate", date))
        chart_pre = _norm(rec.get("ChartPlot"))
        prob_pre = _safe_float(rec.get("Probability"))
        prediction = "Healthy" if failure.lower() == "no failure detected" else "Faulty"
        rpm = _norm(rec.get("RpmGroup"))
        intensity = _norm(rec.get("Probability"))
        intensity_diff = 0

        # Find matching post record
        post = find_post(post_json, tag, date, part, point, failure)
        failure_post = _norm(post.get("FailureDiagnosis")) if post else ""
        prob_post = _safe_float(post.get("Probability")) if post else np.nan
        chart_post = _norm(post.get("ChartPlot")) if post else ""

        # Validation & intensity logic
        failure_val = bool(
            failure_post and failure_post.lower() == failure.lower())

        if np.isnan(prob_pre) and np.isnan(prob_post):
            intensity_val = True
        elif np.isnan(prob_pre) or np.isnan(prob_post):
            intensity_val = False
        else:
            intensity_val = abs(prob_pre - prob_post) <= 15
            intensity_diff = prob_post - prob_pre

        chart_pre_key = _extract_chart_key(chart_pre, machine)
        chart_post_key = _extract_chart_key(chart_post, machine)
        worst_val = chart_pre_key == chart_post_key if chart_post_key else False

        if not failure_val:
            intensity_val = False
            worst_val = False

        # Update or add new row
        found = False
        for r in new_rows:
            same_machine = _norm(r[0]) == machine
            same_part = _norm(r[1]) == part
            same_date = _norm(r[4]) == date_

            # Matching logic for "no failure detected"
            if failure.lower() == "no failure detected" and post:
                if _norm(post.get("FailureDiagnosis", "")) in ['Misalignment', 'Structural looseness', 'Unbalance']:
                    same_point = True
                else:
                    same_point = _norm(r[2]) == _norm(post.get("Point", ""))
            else:
                same_point = _norm(r[2]) == point

            match = (
                same_machine and same_part and same_point and same_date and
                (failure.lower() ==
                 "no failure detected" or _norm(r[5]) == failure)
            )

            if match:
                r[6] = prediction
                r[7] = intensity
                r[8] = failure_val
                r[9] = failure_post
                r[10] = prob_post
                r[11] = _norm(post.get('AIFeedbackNote') if post else "")
                r[12] = intensity_val
                r[14] = intensity_diff
                r[15] = worst_val
                r[17] = chart_post
                found = True
                break

        if not found:
            new_rows.append([
                machine, part, point, rpm, date_, failure,
                prediction, intensity, failure_val, failure_post,
                prob_post, _norm(post.get('AIFeedbackNote') if post else ""),
                intensity_val, "", intensity_diff, worst_val, "", chart_post
            ])

    return new_rows


def _create_dataframe(rows):
    """
    Create a pandas DataFrame with multi-level headers from row data.

    Args:
        rows (list): List of row data.

    Returns:
        pd.DataFrame: Structured DataFrame ready for Excel output.
    """
    header = pd.MultiIndex.from_tuples([
        ("Metadata", "", "Machine"),
        ("Metadata", "", "Part"),
        ("Metadata", "", "Point / Spare"),
        ("Metadata", "", "RPM Group"),
        ("Metadata", "", "Date"),
        ("AI Model", "", "Failure Mode"),
        ("AI Model", "", "Prediction"),
        ("AI Model", "", "Intensity"),
        ("Human Evaluation", "Failure", "Validation"),
        ("Human Evaluation", "Failure", "Failure Mode"),
        ("Human Evaluation", "Failure", "Intensity"),
        ("Human Evaluation", "Failure", "Description"),
        ("Human Evaluation", "Intensity", "Validation"),
        ("Human Evaluation", "Intensity", "Description"),
        ("Human Evaluation", "Intensity", "Intensity Difference"),
        ("Human Evaluation", "Worst Signal", "Validation"),
        ("Human Evaluation", "Worst Signal", "Description"),
        ("Human Evaluation", "", "Worst Signal"),
    ])
    return pd.DataFrame(rows, columns=header)


def _format_excel(output_path):
    """
    Apply formatting to the Excel file: header styles, validation highlighting,
    column widths, and merged machine cells with dashed borders.

    Args:
        output_path (str): Path to the Excel file to format.
    """
    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill(start_color="D1C6F6",
                              end_color="D1C6F6", fill_type="solid")
    invalid_fill = PatternFill(
        start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    dash_border = Border(bottom=Side(border_style="dashed", color="000000"))

    # Format header
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

    # Locate Failure Validation column
    header_row = 3
    failure_val_col = None
    for i, cell in enumerate(ws[header_row], start=1):
        if str(cell.value).strip().lower() == "validation":
            above = ws[header_row - 1][i - 1].value
            if str(above).strip().lower() == "failure":
                failure_val_col = i
                break

    # Highlight invalid rows
    if failure_val_col:
        for row in ws.iter_rows(min_row=header_row + 1):
            if str(row[failure_val_col - 1].value).strip().lower() == "false":
                for cell in row:
                    cell.fill = invalid_fill

    # Summary row
    total_rows = ws.max_row - header_row - 1
    true_count = sum(1 for row in ws.iter_rows(min_row=header_row + 1)
                     if str(row[failure_val_col - 1].value).strip().lower() == "true") if failure_val_col else 0
    percentage_true = (true_count / total_rows * 100) if total_rows > 0 else 0

    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2,
            value=f"Validated Failures: ({percentage_true:.1f}%)").font = Font(bold=True)

    # Auto-fit column widths
    for i, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2

    # Merge Machine cells with dashed borders
    machine_col_idx = None
    for i, cell in enumerate(ws[3], start=1):
        if str(cell.value).strip().lower() == "machine":
            machine_col_idx = i
            break

    if machine_col_idx:
        start_row, last_val = 4, None
        for r in range(start_row, ws.max_row + 2):
            val = ws.cell(
                row=r, column=machine_col_idx).value if r <= ws.max_row else None
            if val != last_val:
                if last_val is not None and r - start_row > 1:
                    ws.merge_cells(start_row=start_row, end_row=r - 1,
                                   start_column=machine_col_idx, end_column=machine_col_idx)
                    merged_cell = ws.cell(
                        row=start_row, column=machine_col_idx)
                    merged_cell.alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=r - 1, column=machine_col_idx).border = dash_border
                start_row, last_val = r, val
            else:
                ws.cell(row=r, column=machine_col_idx).value = None

    wb.save(output_path)
    wb.close()


def build_excel_multilevel(pre_json, post_json, output_path):
    """
    Build a complete Excel report with pre/post AI diagnostics.

    Args:
        pre_json (dict): Pre-test JSON data.
        post_json (dict): Post-test JSON data.
        output_path (str): Path to save the Excel report.

    Returns:
        pd.DataFrame: Final DataFrame used for Excel report.
    """
    rows = []

    for tag, date_dict in pre_json.items():
        for date, records in date_dict.items():
            GetMachineInformation(lst_name=[tag], SelectedDate=date)
            info = Wise(tag)
            new_rows = _generate_base_rows(tag, date, info)
            new_rows = _update_with_records(
                new_rows, records, tag, date, post_json)
            rows.extend(new_rows)

    df = _create_dataframe(rows)
    df.to_excel(output_path)
    _format_excel(output_path)
    return df


def BackTestAuto(company_id, company_domain_before_backtest, company_domain_after_backtest, output_path, domain='20.108.64.195'):
    """
    Automated backtest pipeline: fetch pre/post data, compare diagnostics,
    generate Excel report.

    Args:
        company_id (str): Company identifier.
        company_domain_before_backtest (str): Domain for pre-test data.
        company_domain_after_backtest (str): Domain for post-test data.
        output_path (str): Path to save the final Excel report.
        domain (str, optional): Server IP or domain. Defaults to '20.108.64.195'.
    """
    data_before_backtest = GetCompanyData(
        company_id, domain, company_domain_before_backtest)

    data_after_backtest = GetCompanyData(
        company_id, domain, company_domain_after_backtest)

    if data_before_backtest is None or data_after_backtest is None:
        print("No data fetched. Exiting.")
        return

    build_excel_multilevel(
        pre_json=data_before_backtest,
        post_json=data_after_backtest,
        output_path=output_path
    )
